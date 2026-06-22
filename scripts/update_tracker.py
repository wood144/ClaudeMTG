#!/usr/bin/env python3
"""
Post-game tracker updater for MTG Commander.
Updates Game Log, Rankings, Rankings - Human, Rankings - Claude,
Head-to-Head, and Instructions sheets.

Per-pilot ranking sheets isolate each pilot's W/L with each deck so
deck-quality and pilot-skill variance can be read separately. SoS in
the per-pilot sheets uses the *opposing* pilot's Wilson scores
(opponents in your games were piloted by the other player).

Usage (from mtg-commander/):
    python scripts/update_tracker.py '<JSON game data>'

Example:
    python scripts/update_tracker.py '{
        "date": "2026-03-01",
        "won_by": "Human",
        "winner_deck": "Avatar Allies",
        "loser_deck": "Kellan",
        "win_condition": "Combat Damage",
        "winner_cmdr": "Sokka, Tenacious Tactician",
        "loser_cmdr": "Kellan, the Fae-Blooded",
        "turns": 8,
        "sides": 15,
        "clutch_play": "Standstill locked the door. Appa flew over it.",
        "winner_life": 40,
        "loser_life": 0,
        "comeback": "N",
        "mvp_winner": "Appa, the Vigilant",
        "mvp_loser": "Danitha Capashen",
        "dominance": 2
    }'

Dominance scale (winner only):
    1 = Won because opponent stumbled (mana screw/flood, bad decisions)
    2 = Hard-fought, back and forth
    3 = Deck was humming, gameplan executed
"""

import sys
import json
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TRACKER_PATH = 'assets/mtg_commander_tracker.xlsx'
DECKS_PATH = 'assets/decks.json'
PRICES_CACHE_PATH = 'assets/card_prices.json'
EDHREC_CACHE_PATH = 'assets/edhrec_ranks.json'
Z = 1.645  # 90% CI for Wilson Score
PRICE_CACHE_MAX_AGE = 30  # days — run refresh_prices.py monthly
EDHREC_CACHE_MAX_AGE = 30  # days — run refresh_edhrec.py monthly


# ── Deck name normalization ──────────────────────────────────────────

def load_canonical_deck_names():
    """Load canonical deck names and alias lookup from decks.json."""
    try:
        with open(DECKS_PATH) as f:
            decks = json.load(f)
        if not isinstance(decks, list):
            decks = [{'name': k} for k in decks] if isinstance(decks, dict) else []

        canonical = {}
        alias_map = {}  # lowercase alias -> canonical name
        cmdr_map = {}   # lowercase commander name -> canonical deck name
        for d in decks:
            if 'name' not in d:
                continue
            cname = d['name']
            canonical[cname] = cname
            # Index the canonical name itself (lowercase)
            alias_map[cname.lower()] = cname
            # Index all aliases
            for alias in d.get('aliases', []):
                alias_map[alias.lower()] = cname
            # Index all commanders (handles partners) as lookup keys
            decklist = d.get('list', '')
            if decklist:
                for line in decklist.strip().split('\n'):
                    if '*CMDR*' not in line.upper():
                        # Only first line is commander for non-tagged lists
                        if line == decklist.strip().split('\n')[0]:
                            cmdr_n = line.lstrip('0123456789 ')
                            cmdr_map[cmdr_n.lower()] = cname
                        break
                    cmdr_n = line.lstrip('0123456789 ')
                    # Strip *CMDR* tag
                    cmdr_n = cmdr_n.replace('*CMDR*', '').replace('*cmdr*', '').strip()
                    cmdr_map[cmdr_n.lower()] = cname
        return canonical, alias_map, cmdr_map
    except (FileNotFoundError, json.JSONDecodeError):
        return {}, {}, {}


def load_retired_decks():
    """Return {canonical_name: retired_date_str} for decks flagged retired in
    decks.json. A retired deck is physically disassembled: it keeps its full
    Game Log / Head-to-Head history (and still contributes to opponents' SoS),
    but is shown below a divider in the Rankings rather than ranked as active."""
    try:
        with open(DECKS_PATH) as f:
            decks = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    if not isinstance(decks, list):
        return {}
    return {d['name']: d['retired'] for d in decks
            if d.get('name') and d.get('retired')}


def normalize_deck_name(name, canonical_names, alias_map=None, cmdr_map=None, cmdr_name=None):
    """Match a deck name to its canonical form.
    Checks: commander lookup -> exact match (case-insensitive) -> aliases.
    Falls back to original name if no match found."""
    if alias_map is None:
        alias_map = {}
    if cmdr_map is None:
        cmdr_map = {}

    # Primary: resolve via commander name (most reliable)
    if cmdr_name:
        cmdr_low = cmdr_name.lower().strip()
        match = cmdr_map.get(cmdr_low)
        if not match:
            # Try splitting partner commanders (separated by /, +, or " and ")
            for sep in ['/', '+', ' and ']:
                if sep in cmdr_low:
                    for part in cmdr_low.split(sep):
                        match = cmdr_map.get(part.strip())
                        if match:
                            break
                    if match:
                        break
        if match:
            if match != name:
                print(f"  Resolved deck via commander: '{cmdr_name}' -> '{match}'")
            return match

    low = name.lower().strip()

    # Check alias map (includes canonical names and aliases)
    match = alias_map.get(low)
    if match:
        if match != name:
            print(f"  Normalized deck name: '{name}' -> '{match}'")
        return match

    # Fallback: case-insensitive canonical name match (legacy)
    lower_map = {k.lower(): v for k, v in canonical_names.items()}
    match = lower_map.get(low)
    if match:
        if match != name:
            print(f"  Normalized deck name: '{name}' -> '{match}'")
        return match

    # No match
    print(f"  WARNING: '{name}' not found in decks.json -- using as-is")
    return name


# ── Deck value (from cached Scryfall bulk prices) ────────────────────

def get_deck_values():
    """Return {deck_name: total_usd_value} using cached cheapest-printing prices.
    Cache is built by scripts/refresh_prices.py (run monthly)."""
    try:
        with open(PRICES_CACHE_PATH) as f:
            cache = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        print("  No price cache found. Run: python scripts/refresh_prices.py")
        return {}

    prices = cache.get('prices', {})
    ts = cache.get('timestamp', 'unknown')

    # Warn if stale
    try:
        age_days = (datetime.now() - datetime.fromisoformat(ts)).days
        if age_days > PRICE_CACHE_MAX_AGE:
            print(f"  WARNING: Price cache is {age_days} days old. Run: python scripts/refresh_prices.py")
        else:
            print(f"  Using cached prices ({len(prices)} cards, {age_days}d old, cheapest printing)")
    except (ValueError, TypeError):
        print(f"  Using cached prices ({len(prices)} cards, age unknown)")

    with open(DECKS_PATH) as f:
        decks = json.load(f)

    deck_values = {}
    for d in decks:
        total = 0.0
        for line in d.get('list', '').strip().split('\n'):
            if line.strip():
                parts = line.strip().split(' ', 1)
                qty = int(parts[0]) if parts[0].isdigit() else 1
                name = parts[1] if len(parts) > 1 else parts[0]
                total += qty * prices.get(name, 0.0)
        deck_values[d['name']] = round(total, 2)

    return deck_values


def get_edhrec_ranks():
    """Return {deck_name: edhrec_rank} from cached EDHrec data.
    Cache is built by scripts/refresh_edhrec.py (run monthly)."""
    try:
        with open(EDHREC_CACHE_PATH) as f:
            cache = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        print("  No EDHrec cache found. Run: python scripts/refresh_edhrec.py")
        return {}

    ranks = cache.get('ranks', {})
    ts = cache.get('timestamp', 'unknown')

    try:
        age_days = (datetime.now() - datetime.fromisoformat(ts)).days
        if age_days > EDHREC_CACHE_MAX_AGE:
            print(f"  WARNING: EDHrec cache is {age_days} days old. Run: python scripts/refresh_edhrec.py")
        else:
            print(f"  Using cached EDHrec ranks ({len(ranks)} decks, {age_days}d old)")
    except (ValueError, TypeError):
        print(f"  Using cached EDHrec ranks ({len(ranks)} decks, age unknown)")

    return {name: data.get('rank') for name, data in ranks.items()}


# ── Scoring formulas (from Instructions sheet) ──────────────────────

def wilson_lower(wins, games):
    """Wilson Score lower bound, z=1.645 (90% CI)."""
    if games == 0:
        return 0.0
    p = wins / games
    n = games
    denom = 1 + Z**2 / n
    center = p + Z**2 / (2 * n)
    spread = Z * math.sqrt(p * (1 - p) / n + Z**2 / (4 * n**2))
    return round((center - spread) / denom, 4)


def calc_winner_perf(sides, dominance=2):
    """Winner Perf = 7.0 + speed_bonus + (dominance - 2) * 0.75
    dominance: 1=opponent stumbled, 2=hard-fought, 3=deck humming"""
    speed_bonus = max(0, 1.10 + 0.06 * (17 - sides))
    feels_bonus = (dominance - 2) * 0.75
    return round(7.0 + speed_bonus + feels_bonus, 2)


def calc_loser_perf(sides):
    """Loser Perf = min(1.5, max(0, 2/30 * (sides - 10)))"""
    return round(min(1.5, max(0, (2 / 30) * (sides - 10))), 2)


# ── Rankings helpers (used for overall + per-pilot sheets) ──────────

def aggregate_deck_stats(all_games, pilot=None):
    """
    Build per-deck stats dict. pilot=None for overall (each game contributes
    winner+loser); pilot='human' or 'claude' for per-pilot view (each game
    contributes only the deck that pilot was sleeving).
    """
    deck_stats = {}
    for g in all_games:
        won_by = (g['won_by'] or '').strip().lower()
        if pilot is None:
            entries = [
                (g['winner_deck'], g['winner_cmdr'], True),
                (g['loser_deck'], g['loser_cmdr'], False),
            ]
        elif pilot == 'human':
            if won_by == 'human':
                entries = [(g['winner_deck'], g['winner_cmdr'], True)]
            elif won_by == 'claude':
                entries = [(g['loser_deck'], g['loser_cmdr'], False)]
            else:
                continue
        elif pilot == 'claude':
            if won_by == 'claude':
                entries = [(g['winner_deck'], g['winner_cmdr'], True)]
            elif won_by == 'human':
                entries = [(g['loser_deck'], g['loser_cmdr'], False)]
            else:
                continue
        else:
            raise ValueError(f"Unknown pilot: {pilot}")

        sides = g['sides'] if g['sides'] is not None else 16
        for deck, cmdr, is_winner in entries:
            if deck not in deck_stats:
                deck_stats[deck] = {'cmdr': cmdr, 'wins': 0, 'losses': 0,
                                    'games': 0, 'perfs': [], 'sides_list': []}
            s = deck_stats[deck]
            s['games'] += 1
            if is_winner:
                s['wins'] += 1
                s['perfs'].append(calc_winner_perf(sides, g.get('dominance', 2)))
            else:
                s['losses'] += 1
                s['perfs'].append(calc_loser_perf(sides))
            s['sides_list'].append(sides)
    return deck_stats


def compute_wilson_perf(deck_stats):
    """Compute wilson, perf, avg_sides on each deck stat (in place)."""
    for s in deck_stats.values():
        s['wilson'] = wilson_lower(s['wins'], s['games'])
        s['perf'] = round(sum(s['perfs']) / len(s['perfs']), 2)
        avg_s = sum(s['sides_list']) / len(s['sides_list'])
        s['avg_sides'] = int(avg_s) if avg_s == int(avg_s) else round(avg_s, 1)


def compute_sos_rating(deck_stats, opponent_stats, all_games, pilot=None):
    """
    Compute SoS and Rating for each deck (in place).
    SoS = avg Wilson of decks faced, looked up in `opponent_stats` (the
    deck_stats dict from the opposing pilot's perspective, or overall stats
    for the overall sheet).
    """
    deck_opponents = {}
    for g in all_games:
        won_by = (g['won_by'] or '').strip().lower()
        w, l = g['winner_deck'], g['loser_deck']
        if pilot is None:
            deck_opponents.setdefault(w, []).append(l)
            deck_opponents.setdefault(l, []).append(w)
        elif pilot == 'human':
            if won_by == 'human':
                deck_opponents.setdefault(w, []).append(l)
            elif won_by == 'claude':
                deck_opponents.setdefault(l, []).append(w)
        elif pilot == 'claude':
            if won_by == 'claude':
                deck_opponents.setdefault(w, []).append(l)
            elif won_by == 'human':
                deck_opponents.setdefault(l, []).append(w)

    for deck, s in deck_stats.items():
        opps = deck_opponents.get(deck, [])
        opp_wilsons = [opponent_stats[o]['wilson'] for o in opps if o in opponent_stats]
        s['sos'] = round(sum(opp_wilsons) / len(opp_wilsons), 4) if opp_wilsons else 0.0
        s['rating'] = round(s['wilson'] * (1 + s['sos']), 4)


def write_rankings_sheet(wb, sheet_name, deck_stats, latest_decks, deck_values,
                         edhrec_ranks=None, retired_decks=None):
    """Write a Rankings sheet (overall or per-pilot). Creates sheet if missing.
    Preserves existing trend arrows for decks not in latest_decks.
    Retired decks (in retired_decks) are listed below a '— RETIRED —' divider
    with frozen final stats, no rank number, and no trend arrow."""
    retired_decks = retired_decks or {}
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    # Snapshot prior ranks/trends
    old_ranks = {}
    old_trends = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row and row[0] is not None and len(row) > 2 and row[2] is not None:
            old_ranks[row[2]] = row[0]
            old_trends[row[2]] = row[1] if len(row) > 1 else '-'

    sorted_decks = sorted(
        deck_stats.items(),
        key=lambda x: (x[1]['rating'], x[1]['perf']),
        reverse=True,
    )
    active_sorted = [(d, s) for d, s in sorted_decks if d not in retired_decks]
    retired_sorted = [(d, s) for d, s in sorted_decks if d in retired_decks]
    # Only active decks get a rank number; trend is computed against active ranks.
    new_ranks = {deck: i + 1 for i, (deck, _) in enumerate(active_sorted)}

    def trend_arrow(deck):
        if deck not in latest_decks:
            return old_trends.get(deck, '-')
        old = old_ranks.get(deck)
        if old is None:
            return 'NEW'
        new = new_ranks[deck]
        if new < old:
            return 'UP'
        elif new > old:
            return 'DN'
        return '-'

    edhrec_ranks = edhrec_ranks or {}

    # Clear all existing data rows
    for r in range(2, ws.max_row + 1):
        for c in range(1, 15):
            ws.cell(row=r, column=c).value = None

    bold = Font(bold=True)
    center = Alignment(horizontal='center')
    rank_headers = ['Rank', 'Trend', 'Deck', 'Commander', 'W', 'L', 'Games',
                    'Wilson', 'SoS', 'Rating', 'Perf', 'Avg Sides', 'Value', 'EDHrec']
    for col_idx, hdr in enumerate(rank_headers, start=1):
        ws.cell(1, col_idx, value=hdr).font = bold

    def write_deck_row(r, rank, deck, s, trend, label=None):
        if rank is not None:
            ws.cell(r, 1, value=rank).font = bold
            ws.cell(r, 1).alignment = center
        ws.cell(r, 2, value=trend).alignment = center
        if trend in ('UP', 'DN', 'NEW'):
            ws.cell(r, 2).font = bold
        ws.cell(r, 3, value=label or deck)
        ws.cell(r, 4, value=s['cmdr'])
        ws.cell(r, 5, value=s['wins'])
        ws.cell(r, 6, value=s['losses'])
        ws.cell(r, 7, value=s['games'])
        ws.cell(r, 8, value=s['wilson'])
        ws.cell(r, 9, value=s['sos'])
        ws.cell(r, 10, value=s['rating'])
        ws.cell(r, 11, value=s['perf'])
        ws.cell(r, 12, value=s['avg_sides'])
        val = deck_values.get(deck, 0.0)
        ws.cell(r, 13, value=val).number_format = '$#,##0.00'
        edhrec_rank = edhrec_ranks.get(deck)
        if edhrec_rank is not None:
            ws.cell(r, 14, value=edhrec_rank).alignment = center

    r = 2
    for i, (deck, s) in enumerate(active_sorted):
        write_deck_row(r, i + 1, deck, s, trend_arrow(deck))
        r += 1

    if retired_sorted:
        # Divider, then retired decks with frozen stats, no rank, no trend arrow.
        muted = Font(italic=True, color='888888')
        ws.cell(r, 3, value='— RETIRED —').font = Font(bold=True, color='888888')
        r += 1
        for deck, s in retired_sorted:
            label = f'{deck}  (retired {retired_decks[deck]})'
            write_deck_row(r, None, deck, s, '-', label=label)
            ws.cell(r, 3).font = muted
            r += 1

    return len(active_sorted)


def latest_decks_for_pilot(latest_game, pilot):
    """Return the set of decks that count as 'just played' for a pilot's sheet."""
    if not latest_game:
        return set()
    won_by = (latest_game['won_by'] or '').strip().lower()
    if pilot is None:
        return {latest_game['winner_deck'], latest_game['loser_deck']}
    if pilot == 'human':
        return {latest_game['winner_deck'] if won_by == 'human' else latest_game['loser_deck']}
    if pilot == 'claude':
        return {latest_game['winner_deck'] if won_by == 'claude' else latest_game['loser_deck']}
    return set()


def build_all_rankings(wb, all_games, deck_values, edhrec_ranks=None, retired_decks=None):
    """Build/update the three Rankings sheets: overall, Human, Claude.
    Returns (count_overall, count_human, count_claude) of *active* decks.
    Retired decks still aggregate (so opponents' SoS stays accurate) but render
    below a divider."""
    retired_decks = retired_decks or {}
    overall_stats = aggregate_deck_stats(all_games, pilot=None)
    human_stats = aggregate_deck_stats(all_games, pilot='human')
    claude_stats = aggregate_deck_stats(all_games, pilot='claude')

    compute_wilson_perf(overall_stats)
    compute_wilson_perf(human_stats)
    compute_wilson_perf(claude_stats)

    # SoS: overall uses itself; per-pilot uses the *other* pilot's wilsons
    # (opponents in your games were piloted by the other player).
    compute_sos_rating(overall_stats, overall_stats, all_games, pilot=None)
    compute_sos_rating(human_stats, claude_stats, all_games, pilot='human')
    compute_sos_rating(claude_stats, human_stats, all_games, pilot='claude')

    latest = all_games[-1] if all_games else None

    n_o = write_rankings_sheet(wb, 'Rankings', overall_stats,
                               latest_decks_for_pilot(latest, None), deck_values, edhrec_ranks, retired_decks)
    n_h = write_rankings_sheet(wb, 'Rankings - Human', human_stats,
                               latest_decks_for_pilot(latest, 'human'), deck_values, edhrec_ranks, retired_decks)
    n_c = write_rankings_sheet(wb, 'Rankings - Claude', claude_stats,
                               latest_decks_for_pilot(latest, 'claude'), deck_values, edhrec_ranks, retired_decks)
    return n_o, n_h, n_c


# ── Main update logic ───────────────────────────────────────────────

def update_tracker(game_data):
    wb = openpyxl.load_workbook(TRACKER_PATH)
    canonical, alias_map, cmdr_map = load_canonical_deck_names()
    retired_decks = load_retired_decks()
    bold = Font(bold=True)
    center = Alignment(horizontal='center')

    # Normalize deck names — primarily via commander, fallback to deck name
    game_data['winner_deck'] = normalize_deck_name(
        game_data.get('winner_deck', ''), canonical, alias_map, cmdr_map,
        cmdr_name=game_data.get('winner_cmdr'))
    game_data['loser_deck'] = normalize_deck_name(
        game_data.get('loser_deck', ''), canonical, alias_map, cmdr_map,
        cmdr_name=game_data.get('loser_cmdr'))

    # ── 1. GAME LOG ─────────────────────────────────────────────────
    ws_log = wb['Game Log']
    last_id = ws_log.cell(row=ws_log.max_row, column=1).value
    next_id = (last_id or 0) + 1
    new_row = ws_log.max_row + 1

    date_val = game_data['date']
    if isinstance(date_val, str):
        date_val = datetime.strptime(date_val, '%Y-%m-%d')

    log_values = [
        next_id,
        date_val,
        game_data['won_by'].strip().title(),
        game_data['winner_deck'],
        game_data['loser_deck'],
        game_data['win_condition'],
        game_data['winner_cmdr'],
        game_data['loser_cmdr'],
        game_data.get('turns'),
        game_data.get('sides'),
        game_data.get('clutch_play'),
        game_data.get('winner_life'),
        game_data.get('loser_life'),
        game_data.get('comeback', 'N'),
        game_data.get('mvp_winner'),
        game_data.get('mvp_loser'),
        game_data.get('dominance', 2),
    ]

    for col_idx, val in enumerate(log_values, start=1):
        ws_log.cell(row=new_row, column=col_idx, value=val)
    ws_log.cell(row=new_row, column=2).number_format = 'YYYY-MM-DD'

    print(f"Game Log: added game #{next_id} (row {new_row})")

    # ── 2. REBUILD RANKINGS from Game Log ───────────────────────────
    # Collect every game
    all_games = []
    for row in ws_log.iter_rows(min_row=2, max_row=ws_log.max_row, values_only=True):
        if row[0] is None:
            continue
        all_games.append({
            'id': row[0], 'won_by': row[2],
            'winner_deck': normalize_deck_name(row[3], canonical, alias_map, cmdr_map, cmdr_name=row[6]) if row[3] else row[3],
            'loser_deck': normalize_deck_name(row[4], canonical, alias_map, cmdr_map, cmdr_name=row[7]) if row[4] else row[4],
            'winner_cmdr': row[6], 'loser_cmdr': row[7],
            'sides': row[9],
            'dominance': row[16] if len(row) > 16 and row[16] is not None else 2,
        })

    # Fetch deck values once for all three sheets
    try:
        deck_values = get_deck_values()
    except Exception as e:
        print(f"  Deck values unavailable: {e}")
        deck_values = {}

    try:
        edhrec_ranks = get_edhrec_ranks()
    except Exception as e:
        print(f"  EDHrec ranks unavailable: {e}")
        edhrec_ranks = {}

    n_o, n_h, n_c = build_all_rankings(wb, all_games, deck_values, edhrec_ranks, retired_decks)
    print(f"Rankings: {n_o} active decks ranked")
    print(f"Rankings - Human: {n_h} decks ranked")
    print(f"Rankings - Claude: {n_c} decks ranked")

    # `deck_stats` from the overall view is needed for downstream sheets (H2H,
    # Instructions debuted-deck count). Rebuild a thin reference here.
    deck_stats = aggregate_deck_stats(all_games, pilot=None)
    compute_wilson_perf(deck_stats)
    compute_sos_rating(deck_stats, deck_stats, all_games, pilot=None)

    # ── 3. HEAD-TO-HEAD ─────────────────────────────────────────────
    ws_h2h = wb['Head-to-Head']

    # Build matchup dict: (deckA, deckB) → [A's wins vs B, A's losses vs B]
    matchups = {}
    for g in all_games:
        w, l = g['winner_deck'], g['loser_deck']
        matchups.setdefault((w, l), [0, 0])
        matchups.setdefault((l, w), [0, 0])
        matchups[(w, l)][0] += 1
        matchups[(l, w)][1] += 1

    all_deck_names = sorted(deck_stats.keys())

    # Clear sheet
    for r in range(1, ws_h2h.max_row + 2):
        for c in range(1, ws_h2h.max_column + 2):
            ws_h2h.cell(r, c).value = None

    # Write headers
    hdr_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
    ws_h2h.cell(1, 1, value='Deck').font = bold
    ws_h2h.cell(1, 1).fill = hdr_fill
    ws_h2h.column_dimensions['A'].width = 22

    for j, dk in enumerate(all_deck_names):
        col = j + 2
        ws_h2h.cell(1, col, value=dk).font = bold
        ws_h2h.cell(1, col).fill = hdr_fill
        ws_h2h.cell(j + 2, 1, value=dk).font = bold

    # Fill cells
    for i, dk_row in enumerate(all_deck_names):
        for j, dk_col in enumerate(all_deck_names):
            cell = ws_h2h.cell(i + 2, j + 2)
            cell.alignment = center
            if dk_row == dk_col:
                cell.value = '-'
            else:
                key = (dk_row, dk_col)
                rec = matchups.get(key)
                if rec and (rec[0] > 0 or rec[1] > 0):
                    cell.value = f'{rec[0]}-{rec[1]}'
                else:
                    cell.value = None

    print(f"Head-to-Head: {len(all_deck_names)}x{len(all_deck_names)} matrix")

    # ── 4. INSTRUCTIONS — update series tally ───────────────────────
    ws_inst = wb['Instructions']
    claude_w = sum(1 for g in all_games if g['won_by'].strip().lower() == 'claude')
    human_w = sum(1 for g in all_games if g['won_by'].strip().lower() == 'human')
    total = len(all_games)
    debuted = len(deck_stats)

    # Update row 14 (debuted count) and row 15 (series)
    # Preserve undebuted list manually — just update counts
    old_r14 = ws_inst.cell(14, 1).value or ''
    # Try to preserve undebuted names from existing text
    undebuted_part = ''
    if 'Undebuted:' in old_r14:
        undebuted_part = old_r14[old_r14.index('Undebuted:'):]
    else:
        undebuted_part = f'Undebuted: (check decks.json).'

    retired_debuted = sorted(d for d in retired_decks if d in deck_stats)
    active_debuted = debuted - len(retired_debuted)
    retired_note = f' Retired: {", ".join(retired_debuted)}.' if retired_debuted else ''
    ws_inst.cell(14, 1, value=f'Decks debuted: {debuted} (active {active_debuted}, '
                              f'retired {len(retired_debuted)}).{retired_note} {undebuted_part}')
    ws_inst.cell(15, 1, value=f'Series: Claude {claude_w}, Human {human_w}. Games: {total}.')
    ws_inst.cell(17, 1, value='Trend arrows: Only updated for decks that played. Non-playing decks preserve their existing trend.')

    print(f"Instructions: Series Claude {claude_w}, Human {human_w}. {total} games.")

    # ── Save ────────────────────────────────────────────────────────
    wb.save(TRACKER_PATH)
    print(f"\nSaved to {TRACKER_PATH}")


# ── CLI entry point ─────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) >= 2 and sys.argv[1] == '--refresh-values':
        # Rebuild all 3 ranking sheets (with fresh deck values) without adding a game.
        # Note: trend arrows are preserved as-is (no game played, nothing to compare).
        wb = openpyxl.load_workbook(TRACKER_PATH)
        canonical, alias_map, cmdr_map = load_canonical_deck_names()
        retired_decks = load_retired_decks()
        ws_log = wb['Game Log']
        all_games = []
        for row in ws_log.iter_rows(min_row=2, max_row=ws_log.max_row, values_only=True):
            if row[0] is None:
                continue
            all_games.append({
                'id': row[0], 'won_by': row[2],
                'winner_deck': normalize_deck_name(row[3], canonical, alias_map, cmdr_map, cmdr_name=row[6]) if row[3] else row[3],
                'loser_deck': normalize_deck_name(row[4], canonical, alias_map, cmdr_map, cmdr_name=row[7]) if row[4] else row[4],
                'winner_cmdr': row[6], 'loser_cmdr': row[7],
                'sides': row[9],
                'dominance': row[16] if len(row) > 16 and row[16] is not None else 2,
            })
        try:
            deck_values = get_deck_values()
        except Exception as e:
            print(f"  Deck values unavailable: {e}")
            deck_values = {}
        try:
            edhrec_ranks = get_edhrec_ranks()
        except Exception as e:
            print(f"  EDHrec ranks unavailable: {e}")
            edhrec_ranks = {}
        # Build with empty latest_decks → no trend arrows updated, existing trends preserved
        # (trick: pass an empty all_games tail by calling helpers directly).
        overall_stats = aggregate_deck_stats(all_games, pilot=None)
        human_stats = aggregate_deck_stats(all_games, pilot='human')
        claude_stats = aggregate_deck_stats(all_games, pilot='claude')
        for ds in (overall_stats, human_stats, claude_stats):
            compute_wilson_perf(ds)
        compute_sos_rating(overall_stats, overall_stats, all_games, pilot=None)
        compute_sos_rating(human_stats, claude_stats, all_games, pilot='human')
        compute_sos_rating(claude_stats, human_stats, all_games, pilot='claude')
        n_o = write_rankings_sheet(wb, 'Rankings', overall_stats, set(), deck_values, edhrec_ranks, retired_decks)
        n_h = write_rankings_sheet(wb, 'Rankings - Human', human_stats, set(), deck_values, edhrec_ranks, retired_decks)
        n_c = write_rankings_sheet(wb, 'Rankings - Claude', claude_stats, set(), deck_values, edhrec_ranks, retired_decks)
        wb.save(TRACKER_PATH)
        print(f"Refreshed: Rankings ({n_o}), Rankings - Human ({n_h}), Rankings - Claude ({n_c}).")
        sys.exit(0)

    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    raw = sys.argv[1]
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"Invalid JSON: {e}")
        sys.exit(1)

    # Validate required fields
    required = ['date', 'won_by', 'win_condition', 'winner_cmdr', 'loser_cmdr']
    missing = [f for f in required if f not in data]
    if missing:
        print(f"Missing required fields: {missing}")
        sys.exit(1)

    update_tracker(data)
